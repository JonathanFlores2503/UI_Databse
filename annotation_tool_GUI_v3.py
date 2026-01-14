import os
import glob
import time
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import cv2
from PIL import Image, ImageTk

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


VIDEO_EXTS = (".mp4", ".avi", ".mov", ".mkv", ".m4v", ".wmv", ".webm")

# ---------------------------
# FIXED VIDEO DISPLAY SIZE
# ---------------------------
VIDEO_W = 960
VIDEO_H = 540

# ---------------------------
# PLAYBACK SPEED OPTIONS
# ---------------------------
PLAYBACK_SPEEDS = [0.25, 0.50, 1.75, 1.00, 1.25, 1.50, 1.75, 2.00]
DEFAULT_PLAYBACK_SPEED = 1.75


@dataclass
class VideoItem:
    folder_path: str
    video_path: str
    video_stem: str
    excel_path: str


def normalize_name(name: str) -> str:
    """
    Normalize names to improve compatibility between Excel col A and video filenames.
    - strips whitespace
    - removes extension if present
    - lowercases
    """
    if name is None:
        return ""
    s = str(name).strip()
    s = os.path.splitext(s)[0]  # remove .mp4 if Excel mistakenly includes it
    s = s.strip().lower()
    return s


def excel_build_name_to_row_map(excel_path: str) -> Dict[str, int]:
    """
    Builds a map: normalized_video_name -> row_index
    based on column A (starting row 2).
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    name_to_row: Dict[str, int] = {}

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        key = normalize_name(v)
        if key:
            # If duplicates exist, keep the first one
            if key not in name_to_row:
                name_to_row[key] = r
    return name_to_row


def find_child_folders(parent: str) -> List[str]:
    children = []
    for name in sorted(os.listdir(parent)):
        p = os.path.join(parent, name)
        if os.path.isdir(p):
            children.append(p)
    return children


def find_videos_in_folder(folder: str) -> List[str]:
    vids = []
    for ext in VIDEO_EXTS:
        vids.extend(glob.glob(os.path.join(folder, f"*{ext}")))
        vids.extend(glob.glob(os.path.join(folder, f"*{ext.upper()}")))
    return sorted(set(vids))


def find_existing_excel_in_folder(folder: str) -> Optional[str]:
    """
    Do NOT create any new Excel file.
    Return the first .xlsx found in the folder (sorted), or None if none exists.
    """
    existing = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
    if len(existing) == 0:
        return None
    return existing[0]


def excel_find_row_for_video(ws: Worksheet, video_stem: str) -> Optional[int]:
    target = normalize_name(video_stem)
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if normalize_name(v) == target:
            return r
    return None


def excel_read_labels_from_ws(ws: Worksheet, row: int) -> List[str]:
    """
    Column B (2nd) is split (train/test) and must not be modified.
    Labels start in Column C (3rd) onward.
    Reads contiguous non-empty cells from C.
    """
    labels = []
    c = 3
    while True:
        val = ws.cell(row=row, column=c).value
        if val is None:
            break
        if isinstance(val, str) and val.strip() == "":
            break
        labels.append(str(val).strip())
        c += 1
    return labels


def excel_read_all_labels(excel_path: str) -> Dict[str, List[str]]:
    """
    Returns {video_stem: [labels...]} for all rows.
    Labels start at col C.
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    out: Dict[str, List[str]] = {}

    for r in range(2, ws.max_row + 1):
        stem = ws.cell(row=r, column=1).value
        if not isinstance(stem, str) or not stem.strip():
            continue
        stem = stem.strip()
        out[stem] = excel_read_labels_from_ws(ws, r)

    return out


def excel_write_labels(excel_path: str, video_stem: str, labels_in_order: List[str]) -> bool:
    """
    Writes labels starting from column C for an EXISTING row only.

    Rules:
    - Column A contains the video name (stem). We match against it.
    - Column B is split (train/test) and must NOT be modified.
    - Labels start at Column C (3rd).
    - MUST NOT add new rows. If video_stem is not found, no write occurs.

    Returns:
        True if write occurred, False if video_stem not found.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    row = excel_find_row_for_video(ws, video_stem)
    if row is None:
        return False  # Do not add rows.

    # Clear from column C onward (preserve column B)
    max_col = max(ws.max_column, 3 + len(labels_in_order) + 10)
    for c in range(3, max_col + 1):
        ws.cell(row=row, column=c, value=None)

    # Write labels in the user selection order
    for i, lab in enumerate(labels_in_order):
        ws.cell(row=row, column=3 + i, value=lab)

    wb.save(excel_path)
    return True


class VideoAnnotatorApp:
    def __init__(self, root: tk.Tk, parent_folder: str, labels: List[str]):
        self.root = root
        self.root.title("Video Multi-Label Annotator")

        self.labels = labels[:]
        self.items: List[VideoItem] = []
        self.index = 0

        # Playback state
        self.cap: Optional[cv2.VideoCapture] = None
        self.fps = 30.0
        self.frame_count = 0
        self.duration_s = 0.0
        self.current_frame_idx = 0
        self.playing = False
        self.last_tick = time.time()

        # Playback speed
        self.playback_speed: float = float(DEFAULT_PLAYBACK_SPEED)

        # Selection state
        self.selected: Dict[str, bool] = {lab: False for lab in self.labels}
        self.selection_order: List[str] = []  # preserves order of clicks

        # Cache: excel_path -> {stem -> [labels...]}
        self.folder_label_cache: Dict[str, Dict[str, List[str]]] = {}

        # Slider recursion + debounce
        self._slider_internal_update = False
        self._pending_seek_after_id: Optional[str] = None
        self._pending_seek_target: Optional[int] = None

        self._scan_parent(parent_folder)
        self._build_ui()

        if not self.items:
            messagebox.showwarning(
                "No usable videos found",
                "No videos were found with an existing Excel file in their folder.\n"
                "Each child folder must contain videos AND an existing .xlsx file."
            )
        else:
            self._load_video(0)

        self._tick()

    # -------------------- Data / cache --------------------
    def _scan_parent(self, parent_folder: str) -> None:
        child_folders = find_child_folders(parent_folder)
        skipped = []
        not_in_excel_total = 0

        for folder in child_folders:
            vids = find_videos_in_folder(folder)
            if not vids:
                continue

            excel_path = find_existing_excel_in_folder(folder)
            if excel_path is None:
                skipped.append(folder)
                continue

            try:
                self.folder_label_cache[excel_path] = excel_read_all_labels(excel_path)
                name_to_row = excel_build_name_to_row_map(excel_path)
            except Exception:
                skipped.append(folder)
                continue

            for v in vids:
                stem = os.path.splitext(os.path.basename(v))[0]
                if normalize_name(stem) in name_to_row:
                    self.items.append(VideoItem(folder, v, stem, excel_path))
                else:
                    not_in_excel_total += 1

        if skipped:
            msg = "Some folders were skipped because no readable .xlsx file was found:\n\n"
            msg += "\n".join(skipped[:20])
            if len(skipped) > 20:
                msg += f"\n... ({len(skipped) - 20} more)"
            messagebox.showwarning("Skipped folders", msg)

        if not_in_excel_total > 0:
            messagebox.showinfo(
                "Filename compatibility",
                f"{not_in_excel_total} video(s) were ignored because their filename was not found "
                "in Column A of the corresponding folder Excel."
            )

    def _get_saved_labels_cached(self, item: VideoItem) -> List[str]:
        return self.folder_label_cache.get(item.excel_path, {}).get(item.video_stem, [])

    def _set_saved_labels_cached(self, item: VideoItem, labels: List[str]) -> None:
        if item.excel_path not in self.folder_label_cache:
            self.folder_label_cache[item.excel_path] = {}
        self.folder_label_cache[item.excel_path][item.video_stem] = labels[:]

    def _is_labeled(self, item: VideoItem) -> bool:
        return len(self._get_saved_labels_cached(item)) > 0

    def _count_labeled(self) -> Tuple[int, int]:
        labeled = sum(1 for it in self.items if self._is_labeled(it))
        return labeled, len(self.items)

    # -------------------- UI --------------------
    def _build_ui(self) -> None:
        self.root.geometry("1200x860")

        info_frame = ttk.Frame(self.root, padding=8)
        info_frame.pack(side=tk.TOP, fill=tk.X)

        self.lbl_position = ttk.Label(info_frame, text="0 / 0")
        self.lbl_position.pack(side=tk.LEFT)

        self.lbl_progress = ttk.Label(info_frame, text="Labeled: 0 / 0")
        self.lbl_progress.pack(side=tk.LEFT, padx=(12, 0))

        self.lbl_status_indicator = ttk.Label(info_frame, text="Status: —")
        self.lbl_status_indicator.pack(side=tk.LEFT, padx=(12, 0))

        self.lbl_path = ttk.Label(info_frame, text="", wraplength=820)
        self.lbl_path.pack(side=tk.LEFT, padx=12)

        filter_frame = ttk.Frame(self.root, padding=(8, 0, 8, 8))
        filter_frame.pack(side=tk.TOP, fill=tk.X)

        self.var_skip_done = tk.BooleanVar(value=False)
        self.chk_skip_done = ttk.Checkbutton(
            filter_frame,
            text="Skip done (only navigate unlabeled videos)",
            variable=self.var_skip_done,
            command=self.on_toggle_skip_done
        )
        self.chk_skip_done.pack(side=tk.LEFT)

        # ---------------------------
        # Fixed-size video area
        # ---------------------------
        video_outer = ttk.Frame(self.root, padding=8)
        video_outer.pack(side=tk.TOP, fill=tk.X)

        self.video_container = ttk.Frame(video_outer, width=VIDEO_W, height=VIDEO_H)
        self.video_container.pack(side=tk.TOP, anchor="center")
        self.video_container.pack_propagate(False)

        self.video_label = ttk.Label(self.video_container)
        self.video_label.pack(fill=tk.BOTH, expand=True)

        slider_frame = ttk.Frame(self.root, padding=8)
        slider_frame.pack(side=tk.TOP, fill=tk.X)

        self.slider = ttk.Scale(slider_frame, from_=0, to=1, orient=tk.HORIZONTAL, command=self._on_slider_move)
        self.slider.pack(side=tk.TOP, fill=tk.X)

        self.lbl_time = ttk.Label(slider_frame, text="00:00 / 00:00")
        self.lbl_time.pack(side=tk.TOP, anchor="e")

        controls = ttk.Frame(self.root, padding=8)
        controls.pack(side=tk.TOP, fill=tk.X)

        self.btn_prev = ttk.Button(controls, text="Previous", command=lambda: self.navigate(-1))
        self.btn_prev.pack(side=tk.LEFT)

        self.btn_play = ttk.Button(controls, text="Play", command=self.toggle_play)
        self.btn_play.pack(side=tk.LEFT, padx=8)

        self.btn_next = ttk.Button(controls, text="Next", command=lambda: self.navigate(+1))
        self.btn_next.pack(side=tk.LEFT)

        # ---------------------------
        # Playback speed dropdown
        # ---------------------------
        ttk.Label(controls, text="Speed:").pack(side=tk.LEFT, padx=(24, 6))

        self.speed_var = tk.StringVar(value=f"{DEFAULT_PLAYBACK_SPEED:.2f}")
        speed_values = [f"{s:.2f}" for s in PLAYBACK_SPEEDS]

        self.speed_combo = ttk.Combobox(
            controls,
            textvariable=self.speed_var,
            values=speed_values,
            state="readonly",
            width=6
        )
        self.speed_combo.pack(side=tk.LEFT)
        self.speed_combo.bind("<<ComboboxSelected>>", self.on_speed_change)

        labels_frame = ttk.Frame(self.root, padding=8)
        labels_frame.pack(side=tk.TOP, fill=tk.X)

        ttk.Label(labels_frame, text="Labels (multi-select):").pack(side=tk.LEFT, padx=(0, 10))

        self.label_buttons: Dict[str, tk.Button] = {}
        for lab in self.labels:
            b = tk.Button(
                labels_frame,
                text=lab,
                width=14,
                relief=tk.RAISED,
                command=lambda l=lab: self.toggle_label(l)
            )
            b.pack(side=tk.LEFT, padx=4)
            self.label_buttons[lab] = b

        save_frame = ttk.Frame(self.root, padding=8)
        save_frame.pack(side=tk.TOP, fill=tk.X)

        self.btn_save = ttk.Button(save_frame, text="Save Labels", command=self.save_labels)
        self.btn_save.pack(side=tk.RIGHT)

        self.lbl_status = ttk.Label(save_frame, text="")
        self.lbl_status.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self._refresh_progress_label()

    def on_speed_change(self, event=None) -> None:
        try:
            self.playback_speed = float(self.speed_var.get())
        except Exception:
            self.playback_speed = float(DEFAULT_PLAYBACK_SPEED)
            self.speed_var.set(f"{DEFAULT_PLAYBACK_SPEED:.2f}")

        # Reset timing anchor to avoid dt spikes/jumps
        self.last_tick = time.time()
        self._set_status(f"Playback speed set to {self.playback_speed:.2f}x")

    def _refresh_progress_label(self) -> None:
        labeled, total = self._count_labeled()
        self.lbl_progress.configure(text=f"Labeled: {labeled} / {total}")

    def _set_status_indicator(self, is_labeled: bool) -> None:
        if is_labeled:
            self.lbl_status_indicator.configure(text="Status: already labeled", foreground="#1b5e20")
        else:
            self.lbl_status_indicator.configure(text="Status: unlabeled", foreground="#b71c1c")

    def _set_status(self, msg: str) -> None:
        self.lbl_status.configure(text=msg)

    def on_toggle_skip_done(self) -> None:
        if self.var_skip_done.get() and self.items:
            if self._is_labeled(self.items[self.index]):
                nxt = self._find_next_index(self.index, direction=+1, skip_done=True)
                if nxt is None:
                    messagebox.showinfo("Skip done", "All videos appear to be labeled.")
                else:
                    self._load_video(nxt)

    # -------------------- Video handling --------------------
    def _open_capture(self, video_path: str) -> Tuple[cv2.VideoCapture, float, int]:
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            raise RuntimeError(f"Cannot open video: {video_path}")

        fps = cap.get(cv2.CAP_PROP_FPS)
        if fps is None or fps <= 0:
            fps = 30.0

        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        return cap, float(fps), frame_count

    def _load_video(self, idx: int) -> None:
        if not self.items:
            return

        idx = max(0, min(idx, len(self.items) - 1))
        self.index = idx

        self.playing = False
        self.btn_play.configure(text="Play")
        self.last_tick = time.time()

        if self._pending_seek_after_id is not None:
            try:
                self.root.after_cancel(self._pending_seek_after_id)
            except Exception:
                pass
            self._pending_seek_after_id = None
            self._pending_seek_target = None

        if self.cap is not None:
            try:
                self.cap.release()
            except Exception:
                pass
            self.cap = None

        item = self.items[self.index]
        self.lbl_position.configure(text=f"{self.index + 1} / {len(self.items)}")
        self.lbl_path.configure(text=f"{item.folder_path}  |  {os.path.basename(item.video_path)}")

        self._set_status_indicator(self._is_labeled(item))
        self._refresh_progress_label()

        try:
            self.cap, self.fps, self.frame_count = self._open_capture(item.video_path)
        except Exception as e:
            messagebox.showerror("Video error", str(e))
            return

        self.duration_s = (self.frame_count / self.fps) if self.frame_count > 0 else 0.0
        self.current_frame_idx = 0

        max_frame = max(self.frame_count - 1, 1)
        self.slider.configure(from_=0, to=max_frame)

        saved = self._get_saved_labels_cached(item)
        self.selection_order = []
        for lab in saved:
            if lab in self.labels and lab not in self.selection_order:
                self.selection_order.append(lab)

        self.selected = {lab: (lab in self.selection_order) for lab in self.labels}
        self._refresh_label_buttons()

        self._seek_to_frame(0, update_slider=True)
        self._set_status(f"Loaded. Saved labels: {', '.join(saved) if saved else '(none)'}")

    def _refresh_label_buttons(self) -> None:
        for lab, btn in self.label_buttons.items():
            if self.selected.get(lab, False):
                btn.configure(relief=tk.SUNKEN, bg="#cfe8cf")
            else:
                btn.configure(relief=tk.RAISED, bg=self.root.cget("bg"))

    def _format_time(self, seconds: float) -> str:
        seconds = max(0.0, seconds)
        m = int(seconds // 60)
        s = int(seconds % 60)
        return f"{m:02d}:{s:02d}"

    def _update_time_label(self) -> None:
        cur_s = (self.current_frame_idx / self.fps) if self.fps > 0 else 0.0
        self.lbl_time.configure(text=f"{self._format_time(cur_s)} / {self._format_time(self.duration_s)}")

    def _letterbox_to_fixed(self, frame_rgb):
        fh, fw = frame_rgb.shape[:2]
        if fh <= 0 or fw <= 0:
            return None

        scale = min(VIDEO_W / fw, VIDEO_H / fh)
        new_w = max(1, int(fw * scale))
        new_h = max(1, int(fh * scale))

        resized = cv2.resize(frame_rgb, (new_w, new_h), interpolation=cv2.INTER_AREA)

        pad_left = (VIDEO_W - new_w) // 2
        pad_right = VIDEO_W - new_w - pad_left
        pad_top = (VIDEO_H - new_h) // 2
        pad_bottom = VIDEO_H - new_h - pad_top

        letterboxed = cv2.copyMakeBorder(
            resized,
            top=pad_top, bottom=pad_bottom, left=pad_left, right=pad_right,
            borderType=cv2.BORDER_CONSTANT,
            value=(0, 0, 0)
        )
        return letterboxed

    def _read_frame_at_current(self) -> Optional[ImageTk.PhotoImage]:
        if self.cap is None:
            return None

        ret, frame = self.cap.read()
        if not ret or frame is None or frame.size == 0:
            return None

        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        fixed = self._letterbox_to_fixed(frame_rgb)
        if fixed is None:
            return None

        img = Image.fromarray(fixed)
        return ImageTk.PhotoImage(img)

    def _display_current_frame(self) -> None:
        photo = self._read_frame_at_current()
        if photo is None:
            return
        self.video_label.configure(image=photo)
        self.video_label.image = photo

    def _seek_to_frame(self, frame_idx: int, update_slider: bool) -> None:
        if self.cap is None:
            return

        max_valid = max(self.frame_count - 1, 0)
        frame_idx = int(max(0, min(frame_idx, max_valid)))
        self.current_frame_idx = frame_idx

        self.cap.set(cv2.CAP_PROP_POS_FRAMES, self.current_frame_idx)
        self._display_current_frame()

        if update_slider:
            self._slider_internal_update = True
            try:
                self.slider.set(self.current_frame_idx)
            except tk.TclError:
                pass
            finally:
                self._slider_internal_update = False

        self._update_time_label()

    def _apply_debounced_seek(self) -> None:
        self._pending_seek_after_id = None
        target = self._pending_seek_target
        self._pending_seek_target = None
        if target is None:
            return
        self._seek_to_frame(target, update_slider=False)

    def _on_slider_move(self, value: str) -> None:
        if self._slider_internal_update:
            return

        if self.playing:
            self.playing = False
            self.btn_play.configure(text="Play")

        try:
            v = int(float(value))
        except Exception:
            return

        self._pending_seek_target = v
        if self._pending_seek_after_id is not None:
            try:
                self.root.after_cancel(self._pending_seek_after_id)
            except Exception:
                pass

        self._pending_seek_after_id = self.root.after(30, self._apply_debounced_seek)

    def toggle_play(self) -> None:
        if self.cap is None:
            return
        self.playing = not self.playing
        self.btn_play.configure(text="Pause" if self.playing else "Play")
        self.last_tick = time.time()

    # -------------------- Labels + saving --------------------
    def toggle_label(self, lab: str) -> None:
        if lab not in self.selected:
            return

        new_state = not self.selected[lab]
        self.selected[lab] = new_state

        if new_state:
            if lab not in self.selection_order:
                self.selection_order.append(lab)
        else:
            self.selection_order = [x for x in self.selection_order if x != lab]

        self._refresh_label_buttons()

    def get_selected_labels_in_order(self) -> List[str]:
        ordered = [lab for lab in self.selection_order if self.selected.get(lab, False)]
        for lab in self.labels:
            if self.selected.get(lab, False) and lab not in ordered:
                ordered.append(lab)
        return ordered

    def save_labels(self) -> None:
        if not self.items:
            return

        item = self.items[self.index]
        labels_in_order = self.get_selected_labels_in_order()

        try:
            ok = excel_write_labels(item.excel_path, item.video_stem, labels_in_order)
        except Exception as e:
            messagebox.showerror("Save error", str(e))
            return

        if not ok:
            messagebox.showwarning(
                "Not saved",
                f"Video '{item.video_stem}' was NOT found in the first column of:\n\n{item.excel_path}\n\n"
                "No new row was added. Please ensure Column A contains this video name (without extension)."
            )
            self._set_status("Not saved: video not found in Excel (Column A).")
            return

        self._set_saved_labels_cached(item, labels_in_order)
        self._set_status_indicator(self._is_labeled(item))
        self._refresh_progress_label()

        self._set_status(
            f"Saved (labels from column C): {', '.join(labels_in_order) if labels_in_order else '(none)'}"
        )

    # -------------------- Navigation with skip-done filter --------------------
    def _find_next_index(self, start: int, direction: int, skip_done: bool) -> Optional[int]:
        i = start
        while True:
            i += direction
            if i < 0 or i >= len(self.items):
                return None
            if not skip_done:
                return i
            if not self._is_labeled(self.items[i]):
                return i

    def navigate(self, direction: int) -> None:
        if not self.items:
            return

        skip_done = self.var_skip_done.get()
        nxt = self._find_next_index(self.index, direction=direction, skip_done=skip_done)

        if nxt is None:
            messagebox.showinfo(
                "Navigation",
                "No more unlabeled videos in that direction." if skip_done else "No more videos in that direction."
            )
            return

        self._load_video(nxt)

    # -------------------- Main UI loop --------------------
    def _tick(self) -> None:
        if self.playing and self.cap is not None and self.fps > 0:
            now = time.time()
            dt = now - self.last_tick

            effective_fps = self.fps * float(self.playback_speed)
            frames_to_advance = int(dt * effective_fps)

            if frames_to_advance >= 1:
                new_idx = self.current_frame_idx + frames_to_advance
                if self.frame_count > 0 and new_idx >= self.frame_count:
                    new_idx = self.frame_count - 1
                    self.playing = False
                    self.btn_play.configure(text="Play")

                self._seek_to_frame(new_idx, update_slider=True)
                self.last_tick = now

        self.root.after(20, self._tick)


def main():
    root = tk.Tk()

    parent = filedialog.askdirectory(title="Select parent folder containing child folders with videos")
    if not parent:
        return

    labels = ["Robbery", "Stolen", "Fighting", "Shooting", "Post-Event"]

    VideoAnnotatorApp(root, parent_folder=parent, labels=labels)
    root.mainloop()


if __name__ == "__main__":
    main()
